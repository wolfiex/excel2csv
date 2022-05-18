'''
Making proprietary xlsx formats more accessible.

D. Ellis 2022
'''
# pip install openpyxl pandas
from openpyxl import load_workbook
import argparse, pathlib
import pandas as pd

parser = argparse.ArgumentParser(prog = 'excel2csv', description = 'Convert Excel files to CSV files.')
parser.add_argument( 'fin',nargs=1, type=pathlib.Path, help='Input xlsx file')#argparse.FileType('r')
parser.add_argument('-o','--out', nargs='?', type=str, default='./dataOUT',help='Output Directory') #sys.stdout)


args = parser.parse_args()

try:os.mkdir(args.out)
except:pass

src_file = args.fin[0].__str__()

# Load the spreadsheet
wb = load_workbook(filename = src_file)
origin = src_file.split('/')[-1].split('.')[0]

# Get all the sheets
sheets = wb.sheetnames

for sheet in sheets:

    tables = wb[sheet].tables.keys()

    if len(tables):
        for table in wb[sheet].tables.keys():
            
            dummy = wb[sheet].tables[table]

            df = pd.DataFrame(wb[sheet][dummy.ref]).apply(lambda x: [y.internal_value for y in x])

            df = df.set_index(0,inplace=False).rename(columns=df.iloc[0], inplace = False).iloc[1:]

            out = args.out + '/' + origin + '_' + sheet.replace(' ','-') + '_' + table + '.csv'

            df.to_csv(out)

            print('Written: ' + out)

    else:

        import warnings

        warnings.warn("NO TABLES found in sheet: " + sheet)
        warnings.warn("If using a silly format, please label tables accordingly. \n Using default sheet as a table.")

        df = pd.DataFrame(wb[sheet]).apply(lambda x: [y.internal_value for y in x])

        df = df.set_index(0,inplace=False).rename(columns=df.iloc[0], inplace = False).iloc[1:]

        out = args.out + '/' + origin + '_' + sheet.replace(' ','-')  + '.csv'

        df.to_csv(out)

        print('Written: ' + out)



print('Finished: ' + origin)
